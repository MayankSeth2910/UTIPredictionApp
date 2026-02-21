from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import pickle
import numpy as np
import pandas as pd
from sklearn.preprocessing import StandardScaler
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import json

app = Flask(__name__, static_folder='static')
CORS(app)

# ─── Load Models ──────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODELS_DIR = os.path.join(BASE_DIR, 'models')
DATA_FILE = os.path.join(BASE_DIR, 'predictions_log.xlsx')

rf_model = None
lgbm_untuned = None
lgbm_tuned = None
scaler = None

def load_models():
    global rf_model, lgbm_untuned, lgbm_tuned, scaler
    try:
        with open(os.path.join(MODELS_DIR, 'random_forest_model.pkl'), 'rb') as f:
            rf_model = pickle.load(f)
        with open(os.path.join(MODELS_DIR, 'lgbm_untuned_model.pkl'), 'rb') as f:
            lgbm_untuned = pickle.load(f)
        with open(os.path.join(MODELS_DIR, 'lgbm_tuned_model.pkl'), 'rb') as f:
            lgbm_tuned = pickle.load(f)
        with open(os.path.join(MODELS_DIR, 'scaler.pkl'), 'rb') as f:
            scaler = pickle.load(f)
        print("✅ All models loaded successfully.")
    except Exception as e:
        print(f"⚠️ Model load error: {e}")

# ─── Excel Logger ──────────────────────────────────────────────────────────────
EXCEL_HEADERS = [
    'ID', 'Timestamp', 'Age', 'Gender', 'Color', 'Transparency',
    'pH', 'Specific Gravity', 'WBC', 'RBC', 'Glucose', 'Protein',
    'Epithelial Cells', 'Mucous Threads', 'Amorphous Urates', 'Bacteria',
    'Prediction', 'Confidence', 'Feedback'
]

def init_excel():
    if not os.path.exists(DATA_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Predictions"
        ws.append(EXCEL_HEADERS)
        wb.save(DATA_FILE)

def save_to_excel(row_data: dict, feedback=None):
    try:
        wb = load_workbook(DATA_FILE)
        ws = wb['Predictions']
        row_id = ws.max_row  # auto ID
        row = [
            row_id,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            row_data.get('Age'),
            'FEMALE' if row_data.get('Gender') == 0 else 'MALE',
            row_data.get('Color'),
            row_data.get('Transparency_label'),
            row_data.get('pH'),
            row_data.get('Specific Gravity'),
            row_data.get('WBC'),
            row_data.get('RBC'),
            row_data.get('Glucose_label'),
            row_data.get('Protein_label'),
            row_data.get('Epithelial Cells_label'),
            row_data.get('Mucous Threads_label'),
            row_data.get('Amorphous Urates_label'),
            row_data.get('Bacteria_label'),
            row_data.get('prediction'),
            row_data.get('confidence'),
            feedback if feedback is not None else 'Not provided'
        ]
        ws.append(row)
        wb.save(DATA_FILE)
        return row_id
    except Exception as e:
        print(f"Excel save error: {e}")
        return None

# ─── Preprocessing ─────────────────────────────────────────────────────────────
TRANSPARENCY_MAP = {'CLEAR': 0, 'SLIGHTLY HAZY': 1, 'HAZY': 2, 'TURBID': 3, 'CLOUDY': 4}
GLUCOSE_MAP = {'NEGATIVE': 0, 'TRACE': 1, '1+': 2, '2+': 3, '3+': 4, '4+': 5}
PROTEIN_MAP = {'NEGATIVE': 0, 'TRACE': 1, '1+': 2, '2+': 3, '3+': 4}
EPITHELIAL_MAP = {'NONE SEEN': 0, 'RARE': 1, 'OCCASIONAL': 2, 'FEW': 3, 'MODERATE': 4, 'PLENTY': 5, 'LOADED': 6}
MUCOUS_MAP = {'NONE SEEN': 0, 'RARE': 1, 'OCCASIONAL': 2, 'FEW': 3, 'MODERATE': 4, 'PLENTY': 5}
AMORPHOUS_MAP = {'NONE SEEN': 0, 'RARE': 1, 'OCCASIONAL': 2, 'FEW': 3, 'MODERATE': 4, 'PLENTY': 5}
BACTERIA_MAP = {'NONE SEEN': 0, 'RARE': 1, 'OCCASIONAL': 2, 'FEW': 3, 'MODERATE': 4, 'PLENTY': 5, 'LOADED': 6}
GENDER_MAP = {'FEMALE': 0, 'MALE': 1}
# All possible color dummies from training (adjust if different)
COLOR_COLS = ['Color_AMBER', 'Color_BROWN', 'Color_DARK YELLOW', 'Color_LIGHT RED', 'Color_LIGHT YELLOW',
              'Color_RED', 'Color_REDDISH', 'Color_REDDISH YELLOW', 'Color_STRAW', 'Color_YELLOW']
NUMERICAL_COLS = ['Age', 'pH', 'Specific Gravity', 'WBC', 'RBC',
                  'Transparency', 'Glucose', 'Protein', 'Epithelial Cells',
                  'Mucous Threads', 'Amorphous Urates', 'Bacteria']

def preprocess(data: dict):
    """Transform raw form data into model-ready feature vector."""
    color = data['Color'].upper()
    transparency = TRANSPARENCY_MAP[data['Transparency'].upper()]
    glucose = GLUCOSE_MAP[data['Glucose'].upper()]
    protein = PROTEIN_MAP[data['Protein'].upper()]
    epithelial = EPITHELIAL_MAP[data['Epithelial Cells'].upper()]
    mucous = MUCOUS_MAP[data['Mucous Threads'].upper()]
    amorphous = AMORPHOUS_MAP[data['Amorphous Urates'].upper()]
    bacteria = BACTERIA_MAP[data['Bacteria'].upper()]
    gender = GENDER_MAP[data['Gender'].upper()]

    row = {
        'Age': float(data['Age']),
        'Gender': gender,
        'pH': float(data['pH']),
        'Specific Gravity': float(data['Specific Gravity']),
        'WBC': float(data['WBC']),
        'RBC': float(data['RBC']),
        'Transparency': transparency,
        'Glucose': glucose,
        'Protein': protein,
        'Epithelial Cells': epithelial,
        'Mucous Threads': mucous,
        'Amorphous Urates': amorphous,
        'Bacteria': bacteria,
    }
    # One-hot color
    for col in COLOR_COLS:
        row[col] = 1 if col == f'Color_{color}' else 0

    df = pd.DataFrame([row])
    # Scale numerical cols
    df[NUMERICAL_COLS] = scaler.transform(df[NUMERICAL_COLS])
    expected_cols = ([NUMERICAL_COLS[0]] + ['Gender'] + [NUMERICAL_COLS[5]] + [NUMERICAL_COLS[6]] + [NUMERICAL_COLS[7]] + [NUMERICAL_COLS[1]] 
                     + [NUMERICAL_COLS[2]] + [NUMERICAL_COLS[3]] + [NUMERICAL_COLS[4]] + [NUMERICAL_COLS[8]] + [NUMERICAL_COLS[9]] 
                     + [NUMERICAL_COLS[10]] + [NUMERICAL_COLS[11]] + COLOR_COLS)
    df = df[expected_cols]
    return df

def ensemble_predict(df):
    p_rf = rf_model.predict(df)[0]
    p_lgbm_u = lgbm_untuned.predict(df)[0]
    p_lgbm_t = lgbm_tuned.predict(df)[0]
    votes = [p_rf, p_lgbm_u, p_lgbm_t]
    result = 1 if sum(votes) >= 2 else 0

    # Confidence: average probability
    try:
        proba = np.mean([
            rf_model.predict_proba(df)[0][result],
            lgbm_untuned.predict_proba(df)[0][result],
            lgbm_tuned.predict_proba(df)[0][result],
        ])
    except:
        proba = None

    return result, float(proba) if proba else None

# ─── Routes ───────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/predict', methods=['POST'])
def predict():
    try:
        data = request.json
        df = preprocess(data)
        prediction, confidence = ensemble_predict(df)
        label = 'POSITIVE' if prediction == 1 else 'NEGATIVE'

        # Prepare record for saving
        record = {**data,
                  'Transparency_label': data['Transparency'],
                  'Glucose_label': data['Glucose'],
                  'Protein_label': data['Protein'],
                  'Epithelial Cells_label': data['Epithelial Cells'],
                  'Mucous Threads_label': data['Mucous Threads'],
                  'Amorphous Urates_label': data['Amorphous Urates'],
                  'Bacteria_label': data['Bacteria'],
                  'prediction': label,
                  'confidence': round(confidence * 100, 1) if confidence else None}

        row_id = save_to_excel(record)
        return jsonify({'prediction': label, 'confidence': confidence, 'id': row_id})

    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/feedback', methods=['POST'])
def feedback():
    try:
        data = request.json
        row_id = data.get('id')
        feedback_val = data.get('feedback')  # 'yes' or 'no'

        wb = load_workbook(DATA_FILE)
        ws = wb['Predictions']
        feedback_col = EXCEL_HEADERS.index('Feedback') + 1
        for row in ws.iter_rows(min_row=2):
            if row[0].value == row_id:
                row[feedback_col - 1].value = feedback_val
                break
        wb.save(DATA_FILE)
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

if __name__ == '__main__':
    load_models()
    init_excel()
    app.run(debug=True, port=5000)
